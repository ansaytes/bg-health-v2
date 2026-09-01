import openpyxl, sys, os, math, re, json

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'upload', 'Lagging Indicator (1).xlsx')
TAHUN = 2026

EXCLUDED_SHEETS = ['Data Karyawan Sakit', 'Rekapan Karyawan Sakit']

MONTH_COLS = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']

LEADING_ROWS = {
    6:  'Man Power',
    7:  'Man Hours',
    8:  'Total Kunjungan Klinik',
    9:  'Tenaga Kerja Sakit',
    10: 'Total Absensi Sakit',
    11: 'Spell',
    12: 'Penyakit Akibat Kerja',
    13: 'Kejadian Akibat Penyakit Tenaga Kerja',
    14: 'Layak Bekerja',
}

LAGGING_ROWS = {
    16: 'RKK',
    17: 'CMR',
    18: 'MFR',
    19: 'SSR',
    20: 'ASR',
    21: 'FR PAK',
    22: 'KAPTK',
}

def parse_num(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val) if not math.isnan(val) and not math.isinf(val) else 0.0
    s = str(val).strip().replace('%', '').replace(',', '').strip()
    if s == '' or s == '-' or s.lower() == 'null' or s.lower() == 'undefined':
        return 0.0
    try:
        return float(s)
    except:
        return 0.0

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheets = wb.sheetnames
    
    jobsite_sheets = [s for s in sheets if s not in EXCLUDED_SHEETS]
    jobsite_sheets.sort()
    if 'All Site' in jobsite_sheets:
        jobsite_sheets.remove('All Site')
        jobsite_sheets.insert(0, 'All Site')
    
    rows = []
    
    for site_name in jobsite_sheets:
        ws = wb[site_name]
        
        for row_num, indicator_name in LEADING_ROWS.items():
            row_data = {'tahun': TAHUN, 'site': site_name, 'indicator_type': 'leading', 'indicator_name': indicator_name}
            for i, mkey in enumerate(MONTH_COLS):
                col_idx = 4 + i  
                row_data[mkey] = parse_num(None)
            row_data['ytd'] = parse_num(None)
            rows.append(row_data)
        
        for row_num, indicator_name in LAGGING_ROWS.items():
            row_data = {'tahun': TAHUN, 'site': site_name, 'indicator_type': 'lagging', 'indicator_name': indicator_name}
            for i, mkey in enumerate(MONTH_COLS):
                col_idx = 4 + i
                row_data[mkey] = parse_num(None)
            row_data['ytd'] = parse_num(None)
            rows.append(row_data)
    
    wb.close()
    
    wb2 = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    row_idx = 0
    
    for site_name in jobsite_sheets:
        ws = wb2[site_name]
        
        for row_num, indicator_name in LEADING_ROWS.items():
            row_vals = []
            for i in range(20):
                row_vals.append(None)
            
            for i, mkey in enumerate(MONTH_COLS):
                col_idx = 4 + i
                val = None
                for r in ws.iter_rows(min_row=row_num, max_row=row_num, min_col=col_idx, max_col=col_idx, values_only=True):
                    val = r[0]
                rows[row_idx][mkey] = parse_num(val)
            
            for r in ws.iter_rows(min_row=row_num, max_row=row_num, min_col=16, max_col=16, values_only=True):
                rows[row_idx]['ytd'] = parse_num(r[0])
            
            row_idx += 1
        
        for row_num, indicator_name in LAGGING_ROWS.items():
            for i, mkey in enumerate(MONTH_COLS):
                col_idx = 4 + i
                val = None
                for r in ws.iter_rows(min_row=row_num, max_row=row_num, min_col=col_idx, max_col=col_idx, values_only=True):
                    val = r[0]
                rows[row_idx][mkey] = parse_num(val)
            
            for r in ws.iter_rows(min_row=row_num, max_row=row_num, min_col=16, max_col=16, values_only=True):
                rows[row_idx]['ytd'] = parse_num(r[0])
            
            row_idx += 1
    
    wb2.close()
    
    output_sql = os.path.join(os.path.dirname(__file__), '..', 'download', '005_import_lagging_indicators.sql')
    os.makedirs(os.path.dirname(output_sql), exist_ok=True)
    
    with open(output_sql, 'w', encoding='utf-8') as f:
        f.write(f"-- Import data from Lagging Indicator Excel\n")
        f.write(f"-- Generated automatically - {len(rows)} rows\n")
        f.write(f"-- Tahun: {TAHUN}, Sites: {len(jobsite_sheets)}\n\n")
        f.write("INSERT INTO lagging_indicators (tahun, site, indicator_type, indicator_name, jan, feb, mar, apr, may, jun, jul, aug, sep, oct, nov, dec, ytd) VALUES\n")
        
        for i, r in enumerate(rows):
            vals = [str(r['tahun'])]
            vals.append(f"'{r['site'].replace("'", "''")}'")
            vals.append(f"'{r['indicator_type']}'")
            vals.append(f"'{r['indicator_name'].replace("'", "''")}'")
            for mkey in MONTH_COLS:
                vals.append(str(r[mkey]))
            vals.append(str(r['ytd']))
            
            comma = ',' if i < len(rows) - 1 else ';'
            f.write(f"  ({', '.join(vals)}){comma}\n")
    
    print(f"Generated SQL with {len(rows)} rows -> {output_sql}")
    print(f"Sites: {len(jobsite_sheets)}")
    print(f"Indicators per site: {len(LEADING_ROWS)} leading + {len(LAGGING_ROWS)} lagging = {len(LEADING_ROWS) + len(LAGGING_ROWS)}")
    print(f"Expected rows: {len(jobsite_sheets) * (len(LEADING_ROWS) + len(LAGGING_ROWS))}")
    
    output_json = os.path.join(os.path.dirname(__file__), '..', 'download', 'lagging_indicators_data.json')
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)
    print(f"Also exported JSON -> {output_json}")

if __name__ == '__main__':
    main()
