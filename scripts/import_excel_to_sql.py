#!/usr/bin/env python3
"""
Import Lagging Indicator Excel data -> SQL INSERT statements
Reads all jobsite sheets and Data Karyawan Sakit sheet
Generates SQL file for Supabase SQL Editor
"""
import openpyxl
import sys
import os
from datetime import datetime

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'upload', 'Lagging Indicator (1).xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'download', '004_import_data.sql')

EXCLUDED_SHEETS = ['Data Karyawan Sakit', 'Rekapan Karyawan Sakit']
MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# Leading indicators: row -> column name in DB
LEADING_MAP = {
    6: 'man_power',
    7: 'man_hours',
    8: 'kunjungan_klinik',
    9: 'tk_sakit',
    10: 'absensi_sakit',
    11: 'spell',
    12: 'penyakit_akibat_kerja',
    13: 'kejadian_penyakit_tk',
    14: 'layak_bekerja',
}

# Lagging indicators: row -> column name in DB
LAGGING_MAP = {
    16: 'rkk',
    17: 'cmr',
    18: 'mfr',
    19: 'ssr',
    20: 'asr',
    21: 'fr_pak',
    22: 'kaptk',
}

def parse_num(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace('%', '').replace(',', '').replace(' ', '')
    if s == '' or s == '-':
        return 0
    try:
        return float(s)
    except ValueError:
        return 0

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None
    try:
        dt = datetime.strptime(s, '%Y-%m-%d %H:%M:%S')
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        pass
    try:
        dt = datetime.strptime(s, '%Y-%m-%d')
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        return None

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    sql_lines = []
    sql_lines.append('-- Auto-generated from import_excel_to_sql.py')
    sql_lines.append(f'-- Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    sql_lines.append('')
    sql_lines.append('BEGIN;')
    sql_lines.append('')

    indicator_count = 0
    sick_count = 0

    # Process indicator sheets
    all_sheets = [s for s in wb.sheetnames if s not in EXCLUDED_SHEETS]
    all_sheets.sort(key=lambda x: ('AAA' if x == 'All Site' else 'ZZZ') + x)

    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        print(f'Processing: {sheet_name}')

        # Determine year from row 2 (e.g., "PERIODE 2026")
        year = 2026
        cell_b2 = ws.cell(row=2, column=2).value
        if cell_b2:
            import re
            m = re.search(r'(20\d{2})', str(cell_b2))
            if m:
                year = int(m.group(1))

        for month_idx in range(12):
            col = 4 + month_idx  # D=4 for Jan, E=5 for Feb, etc.
            bulan = month_idx + 1

            # Check if this month has data (check man_power)
            mp_val = parse_num(ws.cell(row=6, column=col).value)
            if mp_val == 0 and all(parse_num(ws.cell(row=r, column=col).value) == 0 for r in [7, 8, 9, 10, 11, 16, 17, 18, 19, 20, 21, 22]):
                continue

            fields = {'tahun': year, 'bulan': bulan, 'jobsite': sheet_name}

            for row, col_name in LEADING_MAP.items():
                fields[col_name] = parse_num(ws.cell(row=row, column=col).value)

            for row, col_name in LAGGING_MAP.items():
                fields[col_name] = parse_num(ws.cell(row=row, column=col).value)

            # Build SQL
            cols = []
            vals = []
            for k, v in fields.items():
                if k == 'jobsite':
                    cols.append(k)
                    vals.append(f"'{v.replace(chr(39), chr(39)+chr(39))}'")
                elif isinstance(v, float):
                    cols.append(k)
                    vals.append(str(v))
                else:
                    cols.append(k)
                    vals.append(str(int(v)))

            sql = f"INSERT INTO health_indicators ({', '.join(cols)}) VALUES ({', '.join(vals)}) ON CONFLICT (tahun, bulan, jobsite) DO UPDATE SET "
            update_parts = []
            for k in cols:
                if k in ('tahun', 'bulan', 'jobsite'):
                    continue
                update_parts.append(f"{k} = EXCLUDED.{k}")
            sql += ', '.join(update_parts) + ';'
            sql_lines.append(sql)
            indicator_count += 1

    # Process sick employees
    if 'Data Karyawan Sakit' in wb.sheetnames:
        print('Processing: Data Karyawan Sakit')
        ws = wb['Data Karyawan Sakit']
        for row in range(2, ws.max_row + 1):
            nama = ws.cell(row=row, column=3).value
            if not nama or not str(nama).strip():
                continue

            nik = str(ws.cell(row=row, column=2).value or '').strip()
            nama = str(nama).strip()
            jobsite = str(ws.cell(row=row, column=4).value or '').strip()
            jabatan = str(ws.cell(row=row, column=5).value or '').strip()

            tgl_mulai_a = parse_date(ws.cell(row=row, column=6).value)
            tgl_selesai_a = parse_date(ws.cell(row=row, column=7).value)
            jml_a = int(parse_num(ws.cell(row=row, column=8).value))
            tgl_mulai_b = parse_date(ws.cell(row=row, column=9).value)
            tgl_selesai_b = parse_date(ws.cell(row=row, column=10).value)
            jml_b = int(parse_num(ws.cell(row=row, column=11).value))
            tgl_mulai_c = parse_date(ws.cell(row=row, column=12).value)
            tgl_selesai_c = parse_date(ws.cell(row=row, column=13).value)
            jml_c = int(parse_num(ws.cell(row=row, column=14).value))
            jml_spell = int(parse_num(ws.cell(row=row, column=15).value))

            def esc(s):
                return s.replace(chr(39), chr(39)+chr(39)) if s else ''

            def null_date(d):
                return f"'{d}'" if d else 'NULL'

            sql = f"""INSERT INTO sick_employees (nik, nama, jobsite, jabatan, tanggal_mulai_a, tanggal_selesai_a, jumlah_hari_a, tanggal_mulai_b, tanggal_selesai_b, jumlah_hari_b, tanggal_mulai_c, tanggal_selesai_c, jumlah_hari_c, jumlah_spell)
VALUES ('{esc(nik)}', '{esc(nama)}', '{esc(jobsite)}', '{esc(jabatan)}', {null_date(tgl_mulai_a)}, {null_date(tgl_selesai_a)}, {jml_a}, {null_date(tgl_mulai_b)}, {null_date(tgl_selesai_b)}, {jml_b}, {null_date(tgl_mulai_c)}, {null_date(tgl_selesai_c)}, {jml_c}, {jml_spell});"""
            sql_lines.append(sql)
            sick_count += 1

    sql_lines.append('')
    sql_lines.append('COMMIT;')

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_lines))

    print(f'\nDone! Generated {indicator_count} indicator rows + {sick_count} sick employee rows')
    print(f'Output: {OUTPUT_PATH}')

if __name__ == '__main__':
    main()
